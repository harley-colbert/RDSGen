from __future__ import annotations
from pathlib import Path
from openpyxl import load_workbook, Workbook
from ..domain.models import Inputs, Computation

LINE_ITEMS_SHEET = 'Line Items'

class CostingGenerator:
    def __init__(self, template_path: str):
        self.template_path = template_path

    def generate(self, out_path: Path, inputs: Inputs, comp: Computation):
        if self.template_path:
            wb = load_workbook(self.template_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = LINE_ITEMS_SHEET
            ws.append(['Line Item','Price','Quantity','Extended'])

        def upsert(name: str, price: float, qty: int):
            found_row = None
            for row in range(2, ws.max_row+1):
                if ws.cell(row=row, column=1).value == name:
                    found_row = row
                    break
            if found_row is None:
                found_row = ws.max_row + 1 if ws.max_row > 1 else 2
                ws.cell(row=found_row, column=1).value = name
            ws.cell(row=found_row, column=2).value = round(price, 4)
            ws.cell(row=found_row, column=3).value = qty
            ws.cell(row=found_row, column=4).value = round(price * qty, 2)

        upsert('Base System', inputs.base_price, 1)
        for name, ext in comp.options_breakdown.items():
            qty = comp.options_qty.get(name, 1)
            unit_price = ext / qty if qty else 0.0
            upsert(name, unit_price, qty)

        total_row = ws.max_row + 1
        ws.cell(row=total_row, column=3).value = 'Options Total'
        ws.cell(row=total_row, column=4).value = comp.options_price_total
        total_row += 1
        ws.cell(row=total_row, column=3).value = 'Grand Total'
        ws.cell(row=total_row, column=4).value = comp.total_price

        wb.save(out_path)
