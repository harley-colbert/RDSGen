from __future__ import annotations
from pathlib import Path
from typing import List, Any, Tuple
import datetime as dt

def _col_letter_to_index(col: str) -> int:
    col = col.upper()
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch) - ord('A') + 1)
    return n

def _parse_a1(a1: str) -> Tuple[int,int]:
    # "B3" -> (row=3, col=2)
    m = __import__("re").match(r"^\s*([A-Za-z]+)(\d+)\s*$", a1)
    if not m:
        raise ValueError(f"Invalid A1 address: {a1}")
    col_l, row_s = m.group(1), m.group(2)
    return int(row_s), _col_letter_to_index(col_l)

def _parse_range(rng: str) -> Tuple[int,int,int,int]:
    # "A1:B3" -> (r1,c1,r2,c2) inclusive
    parts = rng.split(":")
    if len(parts) == 1:
        r, c = _parse_a1(parts[0])
        return r, c, r, c
    r1, c1 = _parse_a1(parts[0])
    r2, c2 = _parse_a1(parts[1])
    if r2 < r1 or c2 < c1:
        r1, r2 = min(r1,r2), max(r1,r2)
        c1, c2 = min(c1,c2), max(c1,c2)
    return r1, c1, r2, c2

class ExternalLinks:
    def __init__(self, external_path: str, compat_mode: bool = False):
        self.path = Path(external_path) if external_path else None
        self.compat_mode = compat_mode

    def touch_log(self, dest_log: Path):
        ts = dt.datetime.now().isoformat(timespec='seconds')
        p = str(self.path) if self.path else "<none>"
        dest_log.write_text(f"Checked connection to {p} at {ts}\n", encoding='utf-8')

    # ---- Public API you can call from controllers/services ----
    def read_range(self, sheet: str, a1_range: str) -> List[List[Any]]:
        if not self.path:
            raise RuntimeError("No EXTERNAL_WORKBOOK_PATH configured.")
        suffix = self.path.suffix.lower()
        if suffix == ".xlsb":
            return self._read_xlsb(sheet, a1_range)
        elif suffix in (".xlsx",".xlsm"):
            return self._read_xlsx(sheet, a1_range)
        else:
            raise RuntimeError(f"Unsupported external workbook type: {suffix}")

    def write_range(self, sheet: str, a1_top_left: str, values: List[List[Any]]) -> None:
        if not self.path:
            raise RuntimeError("No EXTERNAL_WORKBOOK_PATH configured.")
        suffix = self.path.suffix.lower()
        if suffix == ".xlsb":
            if not self.compat_mode:
                raise RuntimeError("Writing to .xlsb requires Excel automation. Enable EXCEL_COMPAT_MODE in Settings (Windows only).")
            return self._write_xlsb_via_com(sheet, a1_top_left, values)
        elif suffix in (".xlsx",".xlsm"):
            return self._write_xlsx(sheet, a1_top_left, values)
        else:
            raise RuntimeError(f"Unsupported external workbook type: {suffix}")

    # ---- XLSX/XLSM via openpyxl ----
    def _read_xlsx(self, sheet: str, a1_range: str) -> List[List[Any]]:
        from openpyxl import load_workbook
        r1, c1, r2, c2 = _parse_range(a1_range)
        wb = load_workbook(self.path, data_only=True, keep_vba=self.path.suffix.lower()==".xlsm")
        if sheet not in wb.sheetnames:
            raise RuntimeError(f"Sheet not found: {sheet}")
        ws = wb[sheet]
        out = []
        for row in ws.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2, values_only=True):
            out.append(list(row))
        return out

    def _write_xlsx(self, sheet: str, a1_top_left: str, values: List[List[Any]]) -> None:
        from openpyxl import load_workbook
        r1, c1, r2, c2 = _parse_range(f"{a1_top_left}:{a1_top_left}")
        wb = load_workbook(self.path, keep_vba=self.path.suffix.lower()==".xlsm")
        ws = wb[sheet] if sheet in wb.sheetnames else wb.create_sheet(title=sheet)
        for i, row in enumerate(values, start=r1):
            for j, val in enumerate(row, start=c1):
                ws.cell(row=i, column=j).value = val
        wb.save(self.path)

    # ---- XLSB read via pyxlsb (read-only) ----
    def _read_xlsb(self, sheet: str, a1_range: str) -> List[List[Any]]:
        from pyxlsb import open_workbook
        r1, c1, r2, c2 = _parse_range(a1_range)
        out = []
        with open_workbook(self.path) as wb:
            with wb.get_sheet(sheet) as ws:
                # pyxlsb is 1-based indexing
                for rr in range(r1, r2+1):
                    row_vals = []
                    for cc in range(c1, c2+1):
                        cell = ws.get_cell(rr, cc)
                        row_vals.append(cell.v if cell else None)
                    out.append(row_vals)
        return out

    # ---- XLSB write via Excel COM (Windows only) ----
    def _write_xlsb_via_com(self, sheet: str, a1_top_left: str, values: List[List[Any]]) -> None:
        try:
            import win32com.client as win32
        except Exception as e:
            raise RuntimeError("pywin32 is required for EXCEL_COMPAT_MODE on Windows.") from e
        r1, c1, _, _ = _parse_range(a1_top_left)
        excel = win32.DispatchEx("Excel.Application")
        excel.DisplayAlerts = False
        try:
            wb = excel.Workbooks.Open(str(self.path))
            try:
                try:
                    ws = wb.Worksheets(sheet)
                except Exception:
                    ws = wb.Worksheets.Add()
                    ws.Name = sheet
                # Build a 2D COM SafeArray
                rows = len(values)
                cols = max((len(r) for r in values), default=0)
                # Determine bottom-right address
                import string
                def col_to_letter(n):
                    s=""
                    while n:
                        n, r = divmod(n-1, 26)
                        s = chr(65+r) + s
                    return s
                top = a1_top_left
                bottom = f"{col_to_letter(c1 + cols - 1)}{r1 + rows - 1}"
                rng = ws.Range(top, bottom)
                # normalize rectangular
                rect = [list(r)+[None]*(cols-len(r)) for r in values]
                rng.Value = rect
                wb.Save()
            finally:
                wb.Close(SaveChanges=False)
        finally:
            excel.Quit()
