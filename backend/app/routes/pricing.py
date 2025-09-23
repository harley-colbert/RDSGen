# backend/app/routes/pricing.py
from __future__ import annotations

import os
import logging
from time import time
from threading import Lock
from pathlib import Path
from urllib.parse import urlparse
from typing import Dict, Tuple, Optional, Any

from flask import request, jsonify, current_app

from .blueprint import api_bp
from .deps import settings_mgr
from ..domain.models import Inputs
from ..domain import rules
from ..services.pricing_engine import ExcelPricingEngine  # ExcelPriceList type optional

log = logging.getLogger("RDSGen.routes.pricing")

# ---------------- Cache of workbook costs (base + per-option) ----------------
_cache_lock = Lock()
_price_cache: Dict[str, object] = {
    "key": None,      # workbook path (str)
    "ts": 0.0,        # when loaded (epoch seconds)
    "base": None,     # float base cost
    "items": None,    # dict[str,float] option costs
    "method": None,   # loader strategy (openpyxl/com)
}


# ----------------------------- Helpers ---------------------------------------

def _is_url(p: str) -> bool:
    try:
        u = urlparse(p)
        return u.scheme in ("http", "https")
    except Exception:
        return False


def _excel_mode_enabled(excel_compat_mode) -> bool:
    """
    Backward + forward compatible:
    - old schema: bool (True == enabled)
    - new schema: 'auto' | 'com' | 'openpyxl' (any of these means enabled)
    """
    if isinstance(excel_compat_mode, bool):
        return bool(excel_compat_mode)
    if isinstance(excel_compat_mode, str):
        return excel_compat_mode.strip().lower() in {"auto", "com", "openpyxl"}
    return False


def preload_cost_cache(*, refresh: bool = False) -> Dict[str, object]:
    """Prime the external workbook cache when configuration allows it.

    Returns a payload describing what happened. When Excel compatibility is
    disabled the cache remains untouched but the function completes
    successfully so callers can continue bootstrapping the app.
    """

    settings = settings_mgr.load(refresh=refresh)
    path = (getattr(settings, "EXTERNAL_WORKBOOK_PATH", "") or "").strip()
    excel_enabled = _excel_mode_enabled(getattr(settings, "EXCEL_COMPAT_MODE", False))

    payload: Dict[str, object] = {
        "excel_enabled": excel_enabled,
        "workbook": path,
        "cache_loaded": False,
        "cache_method": None,
    }

    if not excel_enabled:
        return payload

    if not path:
        raise RuntimeError("EXCEL_COMPAT_MODE is ON but EXTERNAL_WORKBOOK_PATH is empty.")

    if not _is_url(path) and not Path(path).exists():
        raise FileNotFoundError(f"Workbook not found: {path}")

    _ensure_cost_cache(path)

    with _cache_lock:
        payload["cache_loaded"] = True
        payload["cache_ts"] = _price_cache["ts"]
        payload["cache_method"] = _price_cache.get("method")

    return payload


def _ensure_cost_cache(path: str) -> None:
    """
    Load costing grid (read-only baseline) if cache is empty or workbook changed.

    Uses ExcelPricingEngine(path, visible=False) and reads a **margin=0.0** price
    list. This returns base COST and per-option COSTS independent of runtime
    margin. Runtime margin is applied later in rules.compute_from_price_list().
    """
    with _cache_lock:
        if _price_cache["key"] == path and _price_cache["base"] is not None:
            return

    base: float | None = None
    items: Dict[str, float] | None = None
    method: str | None = None

    if not _is_url(path):
        suffix = Path(path).suffix.lower()
        fast_loader = None
        fast_method = None

        if suffix == ".xlsb":
            fast_loader = _read_costs_via_pyxlsb
            fast_method = "pyxlsb"
        else:
            fast_loader = _read_costs_via_openpyxl
            fast_method = "openpyxl"

        try:
            base, items = fast_loader(path)
            method = fast_method
        except Exception as exc:  # pragma: no cover - defensive logging
            current_app.logger.warning(
                "Fast workbook load failed (%s); falling back to COM: %s", fast_method, exc
            )

    if base is None or items is None:
        eng = ExcelPricingEngine(path, visible=False)
        # IMPORTANT: call with positional margin only (no invalid kwargs).
        pl = eng.get_price_list_for_margin(0.0)
        base = float(getattr(pl, "base_price", 0.0) or 0.0)
        items = {str(k): float(v) for k, v in (getattr(pl, "items", {}) or {}).items()}
        method = "com"

    with _cache_lock:
        _price_cache["key"] = path
        _price_cache["ts"] = time()
        _price_cache["base"] = base
        _price_cache["items"] = items
        _price_cache["method"] = method

        if _price_cache["base"] is None or _price_cache["items"] is None:
            raise RuntimeError("ExcelPricingEngine returned an unexpected structure (missing base_price/items).")

        current_app.logger.info(
            "cost_cache refreshed path=%s base=%.2f items=%d method=%s",
            path,
            float(_price_cache["base"]),
            len(_price_cache["items"] or {}),
            _price_cache["method"],
        )


def _read_costs_via_openpyxl(path: str) -> Tuple[float, Dict[str, float]]:
    """Fast path: read Summary sheet via openpyxl without spinning up Excel."""

    from openpyxl import load_workbook  # lazy import to avoid startup cost

    wb = load_workbook(filename=path, data_only=True, read_only=True)
    try:
        if ExcelPricingEngine.SUMMARY not in wb.sheetnames:
            raise RuntimeError(
                f"Summary worksheet ({ExcelPricingEngine.SUMMARY}) not found in workbook"
            )

        ws = wb[ExcelPricingEngine.SUMMARY]

        def _num(cell_addr: str) -> float:
            raw = ws[cell_addr].value
            try:
                return float(raw or 0.0)
            except (TypeError, ValueError):
                return 0.0

        base_total = 0.0
        for row in ExcelPricingEngine.BASE_COMPONENT_ROWS:
            base_total += _num(f"J{row}")

        items: Dict[str, float] = {}
        for row, label in ExcelPricingEngine.PRICE_ROW_LABELS.items():
            items[label] = _num(f"J{row}")

        # Mirror ExcelPricingEngine rounding behaviour for consistency
        base_total = round(base_total, 2)
        items = {k: round(v, 2) for k, v in items.items()}

        return base_total, items
    finally:
        try:
            wb.close()
        except Exception:  # pragma: no cover - best effort cleanup
            pass


def _read_costs_via_pyxlsb(path: str) -> Tuple[float, Dict[str, float]]:
    """Fast path for binary workbooks (.xlsb) using pyxlsb."""

    from pyxlsb import open_workbook  # lazy import; heavy dependency

    def _a1_to_row_col(address: str) -> Tuple[int, int]:
        import re

        m = re.fullmatch(r"\s*([A-Za-z]+)(\d+)\s*", address)
        if not m:
            raise ValueError(f"Invalid cell address: {address}")
        col_letters, row_str = m.groups()
        col = 0
        for ch in col_letters.upper():
            col = col * 26 + (ord(ch) - ord("A") + 1)
        return int(row_str), col

    with open_workbook(path) as wb:
        try:
            ws = wb.get_sheet(ExcelPricingEngine.SUMMARY)
        except KeyError as exc:
            raise RuntimeError(
                f"Summary worksheet ({ExcelPricingEngine.SUMMARY}) not found in workbook"
            ) from exc

        def _num(cell_addr: str) -> float:
            row, col = _a1_to_row_col(cell_addr)
            cell = ws.get_cell(row, col)
            value = getattr(cell, "v", None)
            try:
                return float(value or 0.0)
            except (TypeError, ValueError):
                return 0.0

        base_total = 0.0
        for row in ExcelPricingEngine.BASE_COMPONENT_ROWS:
            base_total += _num(f"J{row}")

        items: Dict[str, float] = {}
        for row, label in ExcelPricingEngine.PRICE_ROW_LABELS.items():
            items[label] = _num(f"J{row}")

        base_total = round(base_total, 2)
        items = {k: round(v, 2) for k, v in items.items()}

        return base_total, items


def _get_cached_costs(path: str) -> Tuple[float, Dict[str, float]]:
    _ensure_cost_cache(path)
    with _cache_lock:
        base = float(_price_cache["base"])  # type: ignore[arg-type]
        items = dict(_price_cache["items"] or {})  # type: ignore[assignment]
        return base, items


# --------- Compatibility helper expected by generate.py (re-added) -----------

def _excel_pricing_if_enabled(margin: float) -> Optional[Any]:
    """
    Compatibility shim for generate.py.

    Return a price list object from the external Excel workbook when Excel mode
    is enabled; else None.

    - Allows SharePoint/HTTPS paths (no local exists() check for URLs).
    - For local paths, verifies the file exists.
    - Uses engine.get_price_list_for_margin(margin), returning the workbook-derived
      structure (typically with .base_price and .items).
    """
    s = settings_mgr.load()
    excel_enabled = _excel_mode_enabled(getattr(s, "EXCEL_COMPAT_MODE", False))
    if not excel_enabled:
        return None

    path = (getattr(s, "EXTERNAL_WORKBOOK_PATH", "") or "").strip()
    if not path:
        raise RuntimeError("EXCEL_COMPAT_MODE is ON but EXTERNAL_WORKBOOK_PATH is empty.")

    if not _is_url(path) and not Path(path).exists():
        raise RuntimeError(f"External workbook not found: {path}")

    eng = ExcelPricingEngine(path, visible=False)
    # IMPORTANT: call with positional margin only.
    return eng.get_price_list_for_margin(float(margin))


# ----------------------------- Endpoints -------------------------------------

@api_bp.post("/price", endpoint="price_live")
def price_live():
    """
    Compute live pricing from cached read-only costs + current inputs/margin.

    Request body supports:
    - either top-level fields or { "inputs": { ... } }
    - legacy `margin` (0..1) and/or `margin_pct` (0..95); models keep them in sync
    """
    payload = request.get_json(force=True) or {}
    data = payload.get("inputs", payload)

    # Validate & normalize inputs early
    try:
        inp = Inputs(**data)
    except Exception as e:
        return jsonify({"ok": False, "errors": {"inputs": str(e)}}), 400

    # Domain-level validation (if your rules module adds constraints)
    val_errors = rules.validate(inp)
    if val_errors:
        return jsonify({"ok": False, "errors": val_errors}), 400

    # Settings and guardrails
    s = settings_mgr.load()
    path = (getattr(s, "EXTERNAL_WORKBOOK_PATH", "") or "").strip()
    excel_enabled = _excel_mode_enabled(getattr(s, "EXCEL_COMPAT_MODE", False))

    if not excel_enabled:
        return jsonify({"ok": False, "errors": {"pricing": "Excel compatibility mode is OFF"}}), 400
    if not path:
        return jsonify({"ok": False, "errors": {"pricing": "External workbook path is empty"}}), 400

    # Existence check only for local files; URLs (SharePoint/HTTPS) are allowed
    if not _is_url(path) and not Path(path).exists():
        return jsonify({"ok": False, "errors": {"pricing": f"Workbook not found: {path}"}}), 400

    # Compute using cached baseline + rules
    try:
        base_cost, item_costs = _get_cached_costs(path)
        comp = rules.compute_from_price_list(inp, base_cost, item_costs)
        payload = comp.model_dump() if hasattr(comp, "model_dump") else comp  # support pydantic/BaseModel or dict

        payload["meta"] = {
            "source": "cache_ro",
            "cache_ts": _price_cache["ts"],
            "workbook": path,
        }
        return jsonify({"ok": True, "pricing": payload})
    except Exception as e:
        current_app.logger.exception("Live pricing (cache) failed for %s", path)
        return jsonify({"ok": False, "errors": {"pricing": f"{type(e).__name__}: {e}"}}), 500


@api_bp.post("/price/refresh", endpoint="price_refresh")
def price_refresh():
    """
    Force refresh of the costing cache (read-only baseline).
    Useful after changing settings (workbook path / compat mode).
    """
    s = settings_mgr.load()
    path = (getattr(s, "EXTERNAL_WORKBOOK_PATH", "") or "").strip()
    excel_enabled = _excel_mode_enabled(getattr(s, "EXCEL_COMPAT_MODE", False))

    if not excel_enabled or not path:
        return jsonify({"ok": False, "errors": {"pricing": "Excel compat OFF or path missing"}}), 400

    if not _is_url(path) and not os.path.exists(path):
        return jsonify({"ok": False, "errors": {"pricing": f"Workbook not found: {path}"}}), 400

    try:
        _ensure_cost_cache(path)
        return jsonify({
            "ok": True,
            "base_cost": _price_cache["base"],
            "items": _price_cache["items"],
            "cache_ts": _price_cache["ts"],
            "workbook": path,
        })
    except Exception as e:
        current_app.logger.exception("Price cache refresh failed for %s", path)
        return jsonify({"ok": False, "errors": {"pricing": f"{type(e).__name__}: {e}"}}), 500
